"""Bulk-import accounts from an XLSX template.

The canonical shape — locked with stakeholder on 2026-06-12 from the
"Data for 5 Accounts with Priorities" spreadsheet — has:

  * Row 4 = headers (rows 1-3 may carry a title block; ignored)
  * Row 5+ = data rows
  * 54 known columns

Column → ORM mapping is in COLUMN_MAP. The 27 product flags live in
PRODUCT_KEYS as `(excel_header, snake_case_key)` pairs and land in the
`account_products` relational table, not as columns on `accounts`.

Dedup rule (stakeholder choice 2026-06-12): if a Client Name matches an
existing live account (case-insensitive, trimmed), RENAME the existing
account to "<old_name>_old" and INSERT the new row. The renamed account
keeps its history; the new one starts fresh.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from decimal import Decimal
from io import BytesIO
from typing import Any
from uuid import UUID

from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.account import Account
from app.models.account_product import AccountProduct
from app.routes.accounts import _slugify, _unique_slug  # reuse existing slug helpers


# ============================================================
# Column catalogue
# ============================================================

# (Excel header) → (Account ORM attribute) for scalar fields. Any column
# absent from this map is either a product flag (see PRODUCT_KEYS) or
# legitimately ignored.
COLUMN_MAP: dict[str, str] = {
    "Client Name":             "name",
    "Priority for the Client": "client_priority",
    "Commercial Owner":        "commercial_owner_name",
    "CSM":                     "csm_owner_name",
    "Live.ai Status":          "platform_status",
    "SubscriptionLevel":       "subscription_plan",
    "Category Count":          "category_count",
    "Supplier Count":          "supplier_count",
    "Segment":                 "segment",
    "Sector":                  "sector",
    "Industry":                "industry",
    "Revenue Bucket":          "revenue_bucket",
    "Billing Country":         "country",
    "Billing Region":          "region",
    "Fortune 500 (Y/N)":       "is_fortune_500",
    "Focus Region (Y/N)":      "is_focus_region",
    "Focus Industy (Y/N)":     "is_focus_industry",  # note: typo in source file preserved
    "Current ACV":             "current_acv",
    "Target ACV":              "target_acv",
    "Annual Revenue in USD":   "annual_revenue_text",  # stored as text per stakeholder
    "Renewal Risk":            "renewal_risk",
    "Tier 1/2/3":              "tier",
    "Tier Category":           "account_type",
    "Procurement Maturity":    "procurement_maturity",
    "GenAI Adoption":          "genai_adoption",
    # `Company Headquarters` intentionally NOT mapped — stakeholder
    # collapsed it into Billing Country (the `country` column).
}

# Boolean (Y/N) columns — converted from "Yes"/"No" → True/False.
BOOL_COLUMNS: frozenset[str] = frozenset({
    "Fortune 500 (Y/N)",
    "Focus Region (Y/N)",
    "Focus Industy (Y/N)",
})

# Integer columns — coerced via int(...) with safe fallback.
INT_COLUMNS: frozenset[str] = frozenset({
    "Category Count",
    "Supplier Count",
})

# Numeric columns — Decimal-safe parse.
DECIMAL_COLUMNS: frozenset[str] = frozenset({
    "Current ACV",
    "Target ACV",
})

# Annual Revenue stays text per stakeholder, but we receive raw numbers
# from the spreadsheet (27590000000). Stringify them on the way in.
TEXT_FROM_NUMBER_COLUMNS: frozenset[str] = frozenset({
    "Annual Revenue in USD",
})

# Excel column → product_key for the account_products table.
PRODUCT_KEYS: list[tuple[str, str]] = [
    ("Category Watch",             "category_watch"),
    ("Category Watch2",            "category_watch_2"),
    ("ABI",                        "abi"),
    ("Supplier Discovery",         "supplier_discovery"),
    ("Supplier Monitoring Risk",   "supplier_monitoring_risk"),
    ("Custom Credits",             "custom_credits"),
    ("Thought Leadership",         "thought_leadership"),
    ("DataHub",                    "datahub"),
    ("Inflation Watch GIT",        "inflation_watch_git"),
    ("Cirtuo",                     "cirtuo"),
    ("Nmanu",                      "nnamu"),  # normalize the spelling drift
    ("Upply",                      "upply"),
    ("Alerts",                     "alerts"),
    ("Commodity Forecasting",      "commodity_forecasting"),
    ("Sourcing Optimizer",         "sourcing_optimizer"),
    ("GSA",                        "gsa"),
    ("Spend Analytics",            "spend_analytics"),
    ("Opp Assessment",             "opp_assessment"),
    ("Diverse Supplier Discovery", "diverse_supplier_discovery"),
    ("Hackett",                    "hackett"),
    ("Abi on CoPilot",             "abi_on_copilot"),
    ("ABI on Teams",               "abi_on_teams"),
    ("ABI HITL",                   "abi_hitl"),
    ("ABI Pro Compare Suppliers",  "abi_pro_compare_suppliers"),
    ("ABI Pro Negotiation Prep",   "abi_pro_negotiation_prep"),
    ("Analyst Validation",         "analyst_validation"),
    ("Prism",                      "prism"),
    ("Procurability Reports",      "procurability_reports"),
]

PRODUCT_KEY_BY_HEADER: dict[str, str] = dict(PRODUCT_KEYS)
ALL_PRODUCT_KEYS: list[str] = [pk for _, pk in PRODUCT_KEYS]


# ============================================================
# Coercion helpers
# ============================================================


def _to_bool(v: Any) -> bool | None:
    """Y/N + Yes/No + truthy strings. Blank → None."""
    if v is None:
        return None
    s = str(v).strip().lower()
    if s in ("", "n/a", "na", "-"):
        return None
    if s in ("y", "yes", "true", "1"):
        return True
    if s in ("n", "no", "false", "0"):
        return False
    return None


def _to_int(v: Any) -> int | None:
    if v is None or v == "":
        return None
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None


def _to_decimal(v: Any) -> Decimal | None:
    if v is None or v == "":
        return None
    if isinstance(v, Decimal):
        return v
    try:
        s = str(v).replace("$", "").replace(",", "").strip()
        return Decimal(s) if s else None
    except (TypeError, ValueError, ArithmeticError):
        return None


def _to_text(v: Any) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _norm_choice(v: Any, *, allowed: tuple[str, ...]) -> str | None:
    """Case-insensitive normalise to one of `allowed`. Unknown → None."""
    s = _to_text(v)
    if not s:
        return None
    for a in allowed:
        if s.lower() == a.lower():
            return a
    return None


# ============================================================
# Row → payload
# ============================================================


@dataclass
class ParsedRow:
    """One row of the input file, post-coercion. `errors` collects
    soft errors (unknown enums etc.) for the preview UI; `name` being
    None is a hard rejection."""

    raw_index: int           # excel row number (1-indexed; for error messages)
    name: str | None = None
    fields: dict[str, Any] = field(default_factory=dict)
    products: dict[str, bool | None] = field(default_factory=dict)
    errors: list[str] = field(default_factory=list)


def parse_xlsx(file_bytes: bytes) -> list[ParsedRow]:
    """Parse the XLSX into ParsedRow objects.

    Header row is row 4 by stakeholder convention; data starts row 5.
    Falls back to the first non-blank row if row 4 isn't headers (so
    other future sheets with the same columns but a different preamble
    still work).
    """
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as e:
        raise RuntimeError("openpyxl is required for XLSX import") from e

    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb[wb.sheetnames[0]]

    headers = _read_headers(ws)
    if not headers:
        return []

    rows: list[ParsedRow] = []
    for r in range(headers["data_start_row"], ws.max_row + 1):
        row = ParsedRow(raw_index=r)
        raw: dict[str, Any] = {}
        for col_idx, header in headers["headers"].items():
            raw[header] = ws.cell(row=r, column=col_idx).value
        # Empty row guard.
        if not any(v is not None and str(v).strip() for v in raw.values()):
            continue
        _coerce_row(raw, row)
        rows.append(row)
    return rows


def _read_headers(ws: Any) -> dict[str, Any] | None:
    # Try row 4 first (stakeholder convention).
    for header_row in (4, 1, 2, 3):
        cell_vals = {
            c: ws.cell(row=header_row, column=c).value
            for c in range(1, ws.max_column + 1)
        }
        # A header row has "Client Name" somewhere.
        if any(_to_text(v) == "Client Name" for v in cell_vals.values()):
            return {
                "header_row": header_row,
                "data_start_row": header_row + 1,
                "headers": {c: _to_text(v) for c, v in cell_vals.items() if _to_text(v)},
            }
    return None


def _coerce_row(raw: dict[str, Any], out: ParsedRow) -> None:
    name = _to_text(raw.get("Client Name"))
    if not name:
        out.errors.append("Missing Client Name")
        return
    out.name = name

    for header, attr in COLUMN_MAP.items():
        v = raw.get(header)
        if header in BOOL_COLUMNS:
            coerced: Any = _to_bool(v)
            # Booleans default to False on the model (NOT NULL with default)
            # — so a blank stays False not None.
            out.fields[attr] = bool(coerced) if coerced is not None else False
        elif header in INT_COLUMNS:
            out.fields[attr] = _to_int(v)
        elif header in DECIMAL_COLUMNS:
            out.fields[attr] = _to_decimal(v)
        elif header in TEXT_FROM_NUMBER_COLUMNS:
            t = _to_text(v)
            out.fields[attr] = t
        elif header == "Renewal Risk":
            val = _norm_choice(v, allowed=("Low", "Medium", "High"))
            if v not in (None, "") and not val:
                out.errors.append(f"Renewal Risk '{v}' isn't Low/Medium/High — left blank.")
            out.fields[attr] = val
        elif header == "Live.ai Status":
            val = _norm_choice(v, allowed=("Active", "Inactive"))
            if v not in (None, "") and not val:
                out.errors.append(f"Live.ai Status '{v}' isn't Active/Inactive — left blank.")
            out.fields[attr] = val
        elif header in ("Procurement Maturity", "GenAI Adoption"):
            val = _norm_choice(v, allowed=("Low", "Medium", "High"))
            if v not in (None, "") and not val:
                out.errors.append(f"{header} '{v}' isn't Low/Medium/High — left blank.")
            out.fields[attr] = val
        else:
            out.fields[attr] = _to_text(v)

    # Products
    for header, product_key in PRODUCT_KEYS:
        out.products[product_key] = _to_bool(raw.get(header))


# ============================================================
# Apply to DB
# ============================================================


@dataclass
class ImportResult:
    created: list[dict[str, Any]] = field(default_factory=list)
    renamed: list[dict[str, Any]] = field(default_factory=list)
    skipped: list[dict[str, Any]] = field(default_factory=list)
    errors: list[dict[str, Any]] = field(default_factory=list)


async def apply_import(
    db: AsyncSession,
    rows: list[ParsedRow],
    *,
    actor_id: UUID,
) -> ImportResult:
    """Apply parsed rows. Dedup rule: if name matches an existing
    non-deleted account (case-insensitive trim), rename existing to
    '<name>_old' and create a fresh row with the incoming data.
    """
    result = ImportResult()

    for row in rows:
        if not row.name:
            result.errors.append({
                "row": row.raw_index,
                "name": None,
                "reason": "; ".join(row.errors) or "Missing name",
            })
            continue

        norm = row.name.strip()

        # Dedup — case-insensitive trim against live accounts.
        existing = (
            await db.execute(
                select(Account).where(
                    func.lower(func.trim(Account.name)) == norm.lower(),
                    Account.deleted_at.is_(None),
                )
            )
        ).scalar_one_or_none()

        if existing is not None:
            # Rename old → "<name>_old" (uniqueify if "_old" already taken).
            existing.name = await _unique_old_name(db, existing.name)
            renamed_info = {
                "row": row.raw_index,
                "name": norm,
                "old_renamed_to": existing.name,
                "old_account_id": str(existing.id),
            }
            result.renamed.append(renamed_info)

        # Build the new account row.
        slug = await _unique_slug(db, _slugify(norm))
        new_acc = Account(name=norm, slug=slug)
        for attr, val in row.fields.items():
            setattr(new_acc, attr, val)
        # Hard default for the boolean-not-null columns when source was blank.
        for b in ("is_fortune_500", "is_focus_region", "is_focus_industry"):
            if getattr(new_acc, b) is None:
                setattr(new_acc, b, False)
        # Default platform status to a known state if we couldn't parse.
        new_acc.cs_entry_type = "A"  # Match the default-on-create rule we already have.
        db.add(new_acc)
        await db.flush()  # need id for account_products inserts

        # Product roster.
        for product_key, purchased in row.products.items():
            db.add(AccountProduct(
                account_id=new_acc.id,
                product_key=product_key,
                purchased=purchased,
                source="import",
                updated_by=actor_id,
            ))

        result.created.append({
            "row": row.raw_index,
            "name": norm,
            "account_id": str(new_acc.id),
            "slug": slug,
            "errors": row.errors,
        })

    return result


async def _unique_old_name(db: AsyncSession, name: str) -> str:
    """'foo' → 'foo_old'; if taken, 'foo_old_2', 'foo_old_3', ..."""
    base = f"{name}_old"
    candidate = base
    n = 1
    while True:
        clash = (
            await db.execute(
                select(Account.id).where(
                    func.lower(Account.name) == candidate.lower(),
                    Account.deleted_at.is_(None),
                )
            )
        ).first()
        if not clash:
            return candidate
        n += 1
        candidate = f"{base}_{n}"
